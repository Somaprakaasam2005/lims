"""
app.py - Minimal Flask application for Civil Engineering LIMS

This is intentionally simple and well-commented for demonstration and learning.
"""
import os
import pkgutil
# Some Python distributions (or newer interpreters) may not provide pkgutil.get_loader
# which Flask's package-finding utilities expect. Provide a small shim if missing.
if not hasattr(pkgutil, 'get_loader'):
    import importlib.util
    def _get_loader(name):
        try:
            # importlib.util.find_spec can raise ValueError in some __main__ contexts
            if name in (None, '__main__'):
                return None
            spec = importlib.util.find_spec(name)
            return spec.loader if spec else None
        except Exception:
            return None
    pkgutil.get_loader = _get_loader

from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
from functools import wraps

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URI', 'sqlite:///lims_dev.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# CSRF Protection Configuration
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_TIME_LIMIT'] = None  # No time limit for CSRF tokens
app.config['SESSION_COOKIE_SECURE'] = False  # Set True in production with HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Initialize CSRF protection
try:
    from flask_wtf.csrf import CSRFProtect
    csrf = CSRFProtect(app)
except ImportError:
    csrf = None
    print('Warning: flask-wtf not installed, CSRF protection disabled')

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Import local modules (models, calculations, report generator)
import models
from calculations import (compressive_strength_mpa, flexural_strength_mpa, 
                         split_tensile_strength_mpa, water_absorption_percent,
                         cbr_value, proctor_compaction, sieve_analysis_summary, atterberg_limits)
from report_generator import generate_test_report_pdf

# Ensure models are initialized with the SQLAlchemy db instance
models.init_models(db)

@login_manager.user_loader
def load_user(user_id):
    return models.User.query.get(int(user_id))

@app.route('/')
@app.route('/dashboard')
@login_required
def index():
    # Get statistics for dashboard
    total_projects = models.Project.query.count()
    total_samples = models.Sample.query.count()
    total_tests = models.TestResult.query.count()
    pending_tests = models.TestResult.query.filter_by(status='Pending').count()
    approved_tests = models.TestResult.query.filter_by(status='Approved').count()
    
    # Recent samples
    recent_samples = models.Sample.query.order_by(models.Sample.id.desc()).limit(5).all()
    
    # Pending tests for approval (if user is Admin or Engineer)
    pending_approval = []
    if current_user.role in ['Admin', 'Engineer']:
        pending_approval = models.TestResult.query.filter_by(status='Pending').limit(10).all()
    
    return render_template('dashboard.html', 
                         user=current_user,
                         total_projects=total_projects,
                         total_samples=total_samples,
                         total_tests=total_tests,
                         pending_tests=pending_tests,
                         approved_tests=approved_tests,
                         recent_samples=recent_samples,
                         pending_approval=pending_approval)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = models.User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid credentials', 'danger')
    return render_template('login.html')


def role_required(*roles):
    """Decorator to require that the current_user has one of the specified roles.

    Usage: @login_required
           @role_required('Admin', 'Lab Technician')
    """
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user or not getattr(current_user, 'is_authenticated', False):
                return login_manager.unauthorized()
            if current_user.role not in roles:
                flash('Permission denied for this action', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return wrapped
    return decorator

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# --- Project Management ---
@app.route('/projects')
@login_required
def projects():
    projects = models.Project.query.order_by(models.Project.id.desc()).all()
    return render_template('projects.html', projects=projects)

@app.route('/projects/new', methods=['GET', 'POST'])
@login_required
@role_required('Admin', 'Lab Technician')
def project_new():
    if request.method == 'POST':
        project_code = request.form.get('project_code', '').strip()
        project_name = request.form.get('project_name', '').strip()
        client_name = request.form.get('client_name', '').strip()
        description = request.form.get('description', '').strip()
        
        if not project_code or not project_name:
            flash('Project Code and Name are required', 'danger')
            return render_template('project_new.html')
        
        # Check for duplicate project code
        if models.Project.query.filter_by(project_code=project_code).first():
            flash('Project Code already exists', 'danger')
            return render_template('project_new.html')
        
        proj = models.Project(
            project_code=project_code,
            project_name=project_name,
            client_name=client_name,
            description=description,
            created_at=datetime.utcnow(),
            status='Active'
        )
        db.session.add(proj)
        db.session.commit()
        flash('Project created', 'success')
        return redirect(url_for('projects'))
    return render_template('project_new.html')

@app.route('/projects/<int:project_id>')
@login_required
def project_detail(project_id):
    proj = models.Project.query.get_or_404(project_id)
    samples = models.Sample.query.filter_by(project_id=project_id).all()
    return render_template('project_detail.html', project=proj, samples=samples)

@app.route('/projects/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Admin', 'Lab Technician')
def project_edit(project_id):
    proj = models.Project.query.get_or_404(project_id)
    if request.method == 'POST':
        project_code = request.form.get('project_code', '').strip()
        project_name = request.form.get('project_name', '').strip()
        client_name = request.form.get('client_name', '').strip()
        description = request.form.get('description', '').strip()
        status = request.form.get('status', 'Active').strip()
        
        if not project_code or not project_name:
            flash('Project Code and Name are required', 'danger')
            return render_template('project_edit.html', project=proj)
        
        # Check for duplicate
        exists = models.Project.query.filter(models.Project.project_code == project_code, models.Project.id != proj.id).first()
        if exists:
            flash('Project Code already exists', 'danger')
            return render_template('project_edit.html', project=proj)
        
        proj.project_code = project_code
        proj.project_name = project_name
        proj.client_name = client_name
        proj.description = description
        proj.status = status
        db.session.commit()
        flash('Project updated', 'success')
        return redirect(url_for('project_detail', project_id=proj.id))
    return render_template('project_edit.html', project=proj)

@app.route('/projects/<int:project_id>/delete', methods=['POST'])
@login_required
@role_required('Admin')
def project_delete(project_id):
    proj = models.Project.query.get_or_404(project_id)
    # Check if project has samples
    if proj.samples:
        flash('Cannot delete project with samples. Delete samples first.', 'danger')
        return redirect(url_for('projects'))
    db.session.delete(proj)
    db.session.commit()
    flash('Project deleted', 'success')
    return redirect(url_for('projects'))

# Sample listing and registration
@app.route('/samples')
@login_required
def samples():
    # Get search and filter parameters
    search = request.args.get('search', '').strip()
    project_filter = request.args.get('project', '').strip()
    type_filter = request.args.get('type', '').strip()
    
    query = models.Sample.query
    
    # Apply filters
    if search:
        query = query.filter(models.Sample.sample_id.ilike(f'%{search}%'))
    if project_filter:
        query = query.filter(models.Sample.project_name.ilike(f'%{project_filter}%'))
    if type_filter:
        query = query.filter_by(sample_type=type_filter)
    
    samples = query.order_by(models.Sample.id.desc()).all()
    
    # Get all projects for dropdown
    projects = models.Project.query.all()
    
    return render_template('samples.html', samples=samples, projects=projects,
                         search=search, project_filter=project_filter, type_filter=type_filter)

@app.route('/samples/new', methods=['GET', 'POST'])
@login_required
@role_required('Admin', 'Lab Technician')
def sample_new():
    projects = models.Project.query.all()
    if request.method == 'POST':
        # Basic sample registration with light validation
        sample_id = request.form.get('sample_id', '').strip()
        sample_type = request.form.get('sample_type', '').strip()
        project_id = request.form.get('project_id', '').strip()
        project_name = request.form.get('project_name', '').strip()
        client_name = request.form.get('client_name', '').strip()
        date_collected = request.form.get('date_collected') or datetime.utcnow().date().isoformat()
        allowed_types = ['Concrete', 'Soil', 'Aggregate']
        if not sample_id:
            flash('Sample ID is required', 'danger')
            return render_template('sample_new.html', projects=projects)
        if sample_type not in allowed_types:
            flash(f'Sample type must be one of {allowed_types}', 'danger')
            return render_template('sample_new.html', projects=projects)
        s = models.Sample(
            sample_id=sample_id, 
            sample_type=sample_type, 
            project_id=int(project_id) if project_id else None,
            project_name=project_name,
            client_name=client_name, 
            date_collected=date_collected
        )
        db.session.add(s)
        db.session.commit()
        flash('Sample registered', 'success')
        return redirect(url_for('samples'))
    return render_template('sample_new.html', projects=projects)

@app.route('/samples/<int:sample_id>', methods=['GET', 'POST'])
@login_required
def sample_detail(sample_id):
    s = models.Sample.query.get_or_404(sample_id)
    tests = s.tests
    if request.method == 'POST':
        test_name = request.form.get('test_name', '').strip()
        raw_value = request.form.get('raw_value', '').strip()
        if not test_name:
            flash('Test name is required', 'danger')
            return redirect(url_for('sample_detail', sample_id=sample_id))
        if not raw_value:
            flash('Raw values are required for the test', 'danger')
            return redirect(url_for('sample_detail', sample_id=sample_id))
        # Save raw values as JSON-like string for simplicity
        tr = models.TestResult(sample_id=s.id, test_name=test_name, raw_values=raw_value, date_tested=datetime.utcnow())
        db.session.add(tr)
        db.session.commit()
        flash('Test added', 'success')
        return redirect(url_for('sample_detail', sample_id=sample_id))
    return render_template('sample_detail.html', sample=s, tests=tests)

@app.route('/samples/<int:sample_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Admin', 'Lab Technician')
def sample_edit(sample_id):
    s = models.Sample.query.get_or_404(sample_id)
    if request.method == 'POST':
        sample_id_code = request.form.get('sample_id', '').strip()
        sample_type = request.form.get('sample_type', '').strip()
        project_name = request.form.get('project_name', '').strip()
        client_name = request.form.get('client_name', '').strip()
        date_collected = request.form.get('date_collected', '').strip()
        
        if not sample_id_code or not sample_type:
            flash('Sample ID and Type are required', 'danger')
            return render_template('sample_edit.html', sample=s)
        
        # Check for duplicate sample_id
        exists = models.Sample.query.filter(models.Sample.sample_id == sample_id_code, models.Sample.id != s.id).first()
        if exists:
            flash('Sample ID already exists', 'danger')
            return render_template('sample_edit.html', sample=s)
        
        s.sample_id = sample_id_code
        s.sample_type = sample_type
        s.project_name = project_name
        s.client_name = client_name
        s.date_collected = date_collected or datetime.utcnow().date().isoformat()
        db.session.commit()
        flash('Sample updated', 'success')
        return redirect(url_for('sample_detail', sample_id=s.id))
    return render_template('sample_edit.html', sample=s)

@app.route('/samples/<int:sample_id>/delete', methods=['POST'])
@login_required
@role_required('Admin')
def sample_delete(sample_id):
    s = models.Sample.query.get_or_404(sample_id)
    # Delete related test results and reports first
    for test in s.tests:
        models.Report.query.filter_by(test_result_id=test.id).delete()
        db.session.delete(test)
    db.session.delete(s)
    db.session.commit()
    flash('Sample deleted', 'success')
    return redirect(url_for('samples'))

@app.route('/calculate/compressive/<int:test_id>')
@login_required
@role_required('Admin', 'Lab Technician')
def calculate_compressive(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    # Expect raw_values like "load_kN,area_mm2"
    try:
        parts = tr.raw_values.split(',')
        load_kN = float(parts[0])
        area_mm2 = float(parts[1])
        strength = compressive_strength_mpa(load_kN, area_mm2)
        tr.calculated_result = f"{strength:.3f} MPa"
        db.session.commit()
        flash('Calculated compressive strength', 'success')
    except Exception as e:
        flash(f'Error calculating: {e}', 'danger')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

@app.route('/calculate/flexural/<int:test_id>')
@login_required
@role_required('Admin', 'Lab Technician')
def calculate_flexural(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    # Expect raw_values like "load_kN,length_mm,width_mm,depth_mm"
    try:
        parts = tr.raw_values.split(',')
        load_kN = float(parts[0])
        length_mm = float(parts[1])
        width_mm = float(parts[2])
        depth_mm = float(parts[3])
        strength = flexural_strength_mpa(load_kN, length_mm, width_mm, depth_mm)
        tr.calculated_result = f"{strength:.3f} MPa"
        db.session.commit()
        flash('Calculated flexural strength', 'success')
    except Exception as e:
        flash(f'Error calculating flexural: {e}', 'danger')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

@app.route('/calculate/split_tensile/<int:test_id>')
@login_required
@role_required('Admin', 'Lab Technician')
def calculate_split_tensile(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    # Expect raw_values like "load_kN,length_mm,diameter_mm"
    try:
        parts = tr.raw_values.split(',')
        load_kN = float(parts[0])
        length_mm = float(parts[1])
        diameter_mm = float(parts[2])
        strength = split_tensile_strength_mpa(load_kN, length_mm, diameter_mm)
        tr.calculated_result = f"{strength:.3f} MPa"
        db.session.commit()
        flash('Calculated split tensile strength', 'success')
    except Exception as e:
        flash(f'Error calculating split tensile: {e}', 'danger')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

@app.route('/calculate/water_absorption/<int:test_id>')
@login_required
@role_required('Admin', 'Lab Technician')
def calculate_water_absorption(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    # Expect raw_values like "dry_mass_g,saturated_mass_g"
    try:
        parts = tr.raw_values.split(',')
        dry_mass = float(parts[0])
        saturated_mass = float(parts[1])
        absorption = water_absorption_percent(dry_mass, saturated_mass)
        tr.calculated_result = f"{absorption:.2f}%"
        db.session.commit()
        flash('Calculated water absorption', 'success')
    except Exception as e:
        flash(f'Error calculating water absorption: {e}', 'danger')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

@app.route('/calculate/cbr/<int:test_id>')
@login_required
@role_required('Admin', 'Lab Technician')
def calculate_cbr(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    # Expect raw_values like "load_kN,standard_load_kN"
    try:
        parts = tr.raw_values.split(',')
        load = float(parts[0])
        standard = float(parts[1])
        cbr = cbr_value(load, standard)
        tr.calculated_result = f"CBR = {cbr:.2f}%"
        db.session.commit()
        flash('Calculated CBR value', 'success')
    except Exception as e:
        flash(f'Error calculating CBR: {e}', 'danger')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

@app.route('/calculate/proctor/<int:test_id>')
@login_required
@role_required('Admin', 'Lab Technician')
def calculate_proctor(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    # Expect raw_values like "dry_density_kgm3,water_content_percent"
    try:
        parts = tr.raw_values.split(',')
        dry_density = float(parts[0])
        water_content = float(parts[1])
        result = proctor_compaction(dry_density, water_content)
        tr.calculated_result = f"ρd={result['dry_density']} kg/m³, w={result['water_content']}%"
        db.session.commit()
        flash('Calculated Proctor compaction data', 'success')
    except Exception as e:
        flash(f'Error calculating Proctor: {e}', 'danger')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

@app.route('/calculate/sieve/<int:test_id>')
@login_required
@role_required('Admin', 'Lab Technician')
def calculate_sieve(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    # Expect raw_values as semicolon-separated mass retained per sieve: "sieve:mass;sieve:mass;..." and total mass
    try:
        # Example: "75:10;37.5:20;19:30;9.5:25;4.75:10;total:95"
        entries = [e for e in tr.raw_values.split(';') if e]
        sieve_masses = {}
        total = None
        for e in entries:
            k, v = e.split(':')
            if k.strip().lower() == 'total':
                total = float(v)
            else:
                sieve_masses[float(k.strip())] = float(v)
        if total is None:
            total = sum(sieve_masses.values())
        summary = sieve_analysis_summary(sieve_masses, total)
        tr.calculated_result = str(summary['summary_table'])
        db.session.commit()
        flash('Sieve analysis calculated', 'success')
    except Exception as e:
        flash(f'Error calculating sieve: {e}', 'danger')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

@app.route('/calculate/atterberg/<int:test_id>')
@login_required
@role_required('Admin', 'Lab Technician')
def calculate_atterberg(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    # Expect raw_values like "LL,PL" (liquid limit, plastic limit)
    try:
        parts = tr.raw_values.split(',')
        liquid_limit = float(parts[0])
        plastic_limit = float(parts[1])
        result = atterberg_limits(liquid_limit, plastic_limit)
        tr.calculated_result = f"LL={result['LL']}%, PL={result['PL']}%, PI={result['PI']}%"
        db.session.commit()
        flash('Atterberg limits calculated', 'success')
    except Exception as e:
        flash(f'Error calculating Atterberg: {e}', 'danger')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

# --- Result Approval Workflow ---
@app.route('/test/<int:test_id>/approve', methods=['POST'])
@login_required
@role_required('Admin', 'Engineer')
def approve_test(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    remarks = request.form.get('remarks', '').strip()
    
    tr.status = 'Approved'
    tr.approved_by = current_user.id
    tr.approved_at = datetime.utcnow()
    tr.remarks = remarks
    db.session.commit()
    flash('Test result approved', 'success')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

@app.route('/test/<int:test_id>/reject', methods=['POST'])
@login_required
@role_required('Admin', 'Engineer')
def reject_test(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    remarks = request.form.get('remarks', '').strip()
    
    if not remarks:
        flash('Remarks are required when rejecting a test', 'danger')
        return redirect(url_for('sample_detail', sample_id=tr.sample_id))
    
    tr.status = 'Rejected'
    tr.approved_by = current_user.id
    tr.approved_at = datetime.utcnow()
    tr.remarks = remarks
    db.session.commit()
    flash('Test result rejected', 'danger')
    return redirect(url_for('sample_detail', sample_id=tr.sample_id))

@app.route('/reports/generate/<int:test_id>')
@login_required
@role_required('Admin', 'Lab Technician', 'Engineer')
def generate_report(test_id):
    tr = models.TestResult.query.get_or_404(test_id)
    sample = tr.sample
    lab_name = 'Civil Engg Materials Lab - College'
    out_path = f"reports/report_{tr.id}.pdf"
    os.makedirs('reports', exist_ok=True)
    generate_test_report_pdf(out_path, lab_name=lab_name, sample=sample, test_name=tr.test_name,
                             raw_values=tr.raw_values, result=tr.calculated_result, technician=current_user.username)
    # Save a record
    rpt = models.Report(sample_id=sample.id, test_result_id=tr.id, file_path=out_path, created_at=datetime.utcnow())
    db.session.add(rpt)
    db.session.commit()
    
    # Audit log
    log_audit('GENERATE_REPORT', 'TestResult', tr.id, f'Generated PDF report for test {tr.test_name}')
    
    flash('Report generated', 'success')
    return send_file(out_path, as_attachment=True)

@app.route('/reports/batch', methods=['POST'])
@login_required
@role_required('Admin', 'Lab Technician', 'Engineer')
def generate_batch_report():
    """Generate batch PDF report for multiple tests"""
    test_ids = request.form.getlist('test_ids')
    if not test_ids:
        flash('No tests selected for batch report', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    
    tests = models.TestResult.query.filter(models.TestResult.id.in_(test_ids)).all()
    if not tests:
        flash('No valid tests found', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    
    # Group by sample
    sample = tests[0].sample
    lab_name = 'Civil Engg Materials Lab - College'
    out_path = f"reports/batch_{sample.id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.pdf"
    os.makedirs('reports', exist_ok=True)
    
    # Generate combined PDF with all tests
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    
    c = canvas.Canvas(out_path, pagesize=A4)
    width, height = A4
    
    # Title page
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(width/2, height-2*inch, lab_name)
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width/2, height-2.5*inch, "Batch Test Report")
    c.setFont("Helvetica", 12)
    c.drawCentredString(width/2, height-3*inch, f"Sample: {sample.sample_id}")
    c.drawCentredString(width/2, height-3.3*inch, f"Project: {sample.project.project_name}")
    c.drawCentredString(width/2, height-3.6*inch, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}")
    c.drawCentredString(width/2, height-3.9*inch, f"Total Tests: {len(tests)}")
    
    # New page for each test
    for test in tests:
        c.showPage()
        y = height - inch
        c.setFont("Helvetica-Bold", 14)
        c.drawString(inch, y, f"Test: {test.test_name}")
        y -= 0.3*inch
        c.setFont("Helvetica", 11)
        c.drawString(inch, y, f"Status: {test.status}")
        y -= 0.3*inch
        c.drawString(inch, y, f"Raw Values: {test.raw_values}")
        y -= 0.3*inch
        c.drawString(inch, y, f"Result: {test.calculated_result}")
        y -= 0.3*inch
        c.drawString(inch, y, f"Tested: {test.created_at.strftime('%Y-%m-%d')}")
        if test.approved_at:
            y -= 0.3*inch
            c.drawString(inch, y, f"Approved: {test.approved_at.strftime('%Y-%m-%d')}")
        if test.remarks:
            y -= 0.3*inch
            c.drawString(inch, y, f"Remarks: {test.remarks}")
    
    c.save()
    
    # Save report record
    rpt = models.Report(sample_id=sample.id, test_result_id=None, file_path=out_path, created_at=datetime.utcnow())
    db.session.add(rpt)
    db.session.commit()
    
    log_audit('GENERATE_BATCH_REPORT', 'Sample', sample.id, f'Generated batch report for {len(tests)} tests')
    
    flash(f'Batch report generated for {len(tests)} tests', 'success')
    return send_file(out_path, as_attachment=True, download_name=f'batch_{sample.sample_id}.pdf')

# --- Excel Export ---
@app.route('/export/samples')
@login_required
def export_samples():
    """Export all samples to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Samples"
        
        # Header
        headers = ['ID', 'Sample ID', 'Type', 'Project', 'Client', 'Date Collected', 'Test Count']
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data
        samples = models.Sample.query.all()
        for s in samples:
            project_name = s.project.project_name if s.project else (s.project_name or 'N/A')
            ws.append([s.id, s.sample_id, s.sample_type, project_name, 
                      s.client_name, s.date_collected, len(s.tests)])
        
        # Save
        output_path = 'reports/samples_export.xlsx'
        wb.save(output_path)
        
        log_audit('EXPORT', 'Sample', None, f'Exported {len(samples)} samples to Excel')
        
        return send_file(output_path, as_attachment=True, download_name='samples.xlsx')
    except ImportError:
        flash('openpyxl not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('samples'))
    except Exception as e:
        flash(f'Error exporting: {e}', 'danger')
        return redirect(url_for('samples'))

@app.route('/export/tests')
@login_required
def export_tests():
    """Export all test results to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        
        # Header
        headers = ['Test ID', 'Sample ID', 'Test Name', 'Raw Values', 'Result', 'Status', 'Date Tested', 'Approved By']
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data
        tests = models.TestResult.query.all()
        for t in tests:
            approved_by = t.approver.username if t.approver else 'N/A'
            ws.append([t.id, t.sample.sample_id, t.test_name, t.raw_values, 
                      t.calculated_result, t.status or 'Pending', 
                      t.date_tested.strftime('%Y-%m-%d') if t.date_tested else '', approved_by])
        
        # Save
        output_path = 'reports/tests_export.xlsx'
        wb.save(output_path)
        
        log_audit('EXPORT', 'TestResult', None, f'Exported {len(tests)} tests to Excel')
        
        return send_file(output_path, as_attachment=True, download_name='tests.xlsx')
    except ImportError:
        flash('openpyxl not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('samples'))
    except Exception as e:
        flash(f'Error exporting: {e}', 'danger')
        return redirect(url_for('samples'))

# --- Audit Log ---
def log_audit(action, entity_type, entity_id, details):
    """Helper function to create audit log entries."""
    try:
        log = models.AuditLog(
            user_id=current_user.id if current_user.is_authenticated else None,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            details=details,
            timestamp=datetime.utcnow()
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        pass  # Don't fail operations if audit logging fails

@app.route('/audit/logs')
@login_required
@role_required('Admin')
def audit_logs():
    """View audit logs."""
    logs = models.AuditLog.query.order_by(models.AuditLog.timestamp.desc()).limit(100).all()
    return render_template('audit_logs.html', logs=logs)


# --- User management (Admin only) ---
@app.route('/users')
@login_required
@role_required('Admin')
def users():
    users = models.User.query.order_by(models.User.id.desc()).all()
    return render_template('users.html', users=users)


@app.route('/users/new', methods=['GET', 'POST'])
@login_required
@role_required('Admin')
def user_new():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'Lab Technician').strip()
        if not username or not password:
            flash('Username and password are required', 'danger')
            return render_template('user_new.html')
        # prevent duplicate usernames
        if models.User.query.filter_by(username=username).first():
            flash('Username already exists', 'danger')
            return render_template('user_new.html')
        u = models.User(username=username, role=role)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        flash('User created', 'success')
        return redirect(url_for('users'))
    return render_template('user_new.html')


@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Admin')
def user_edit(user_id):
    u = models.User.query.get_or_404(user_id)
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        role = request.form.get('role', '').strip()
        if not username:
            flash('Username is required', 'danger')
            return render_template('user_edit.html', user=u)
        # prevent collisions
        exists = models.User.query.filter(models.User.username == username, models.User.id != u.id).first()
        if exists:
            flash('Username already taken', 'danger')
            return render_template('user_edit.html', user=u)
        u.username = username
        u.role = role
        if request.form.get('password'):
            u.set_password(request.form.get('password'))
        db.session.commit()
        flash('User updated', 'success')
        return redirect(url_for('users'))
    return render_template('user_edit.html', user=u)


@app.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@role_required('Admin')
def user_delete(user_id):
    u = models.User.query.get_or_404(user_id)
    # prevent deleting self
    if u.id == current_user.id:
        flash('Cannot delete yourself', 'danger')
        return redirect(url_for('users'))
    db.session.delete(u)
    db.session.commit()
    flash('User deleted', 'success')
    return redirect(url_for('users'))

# Error handlers
@app.errorhandler(404)
def not_found(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()  # Rollback any failed database operations
    return render_template('500.html'), 500

if __name__ == '__main__':
    # Ensure we run DB setup inside the app context
    with app.app_context():
        db.create_all()
        # Create an admin user for quick testing if none exists
        try:
            if models.User.query.count() == 0:
                admin = models.User(username='admin', role='Admin')
                admin.set_password('admin')
                db.session.add(admin)
                db.session.commit()
                print('Created default admin user: admin / admin')
        except Exception:
            # If models or DB are not yet fully configured, skip admin creation
            pass
    
    # Use environment variables for debug mode
    debug_mode = os.getenv('FLASK_DEBUG', 'False').lower() in ('true', '1', 'yes')
    app.run(debug=debug_mode)
